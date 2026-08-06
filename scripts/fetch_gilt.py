import io
import json
import os
import zipfile
from datetime import datetime

import requests
from openpyxl import load_workbook

URL = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip"
HISTORY_FILE = "data/gilt-2y.json"
MAX_DAYS = 16

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}


def load_history():
    if not os.path.exists(HISTORY_FILE):
        return []
    try:
        with open(HISTORY_FILE) as f:
            data = json.load(f)
        # Support old single-object format
        if isinstance(data, dict):
            return [data]
        return data
    except Exception:
        return []


def save_history(history):
    os.makedirs("data", exist_ok=True)
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2)


def fetch_latest_from_boe():
    resp = requests.get(URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()

    with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
        xlsx_name = next(
            n for n in z.namelist()
            if n.lower().endswith(".xlsx") and "nominal" in n.lower()
        )
        with z.open(xlsx_name) as f:
            wb = load_workbook(io.BytesIO(f.read()), data_only=True)

    sheet = wb["4. spot curve"]

    # Find the 2-year column
    col_2y = None
    for c in range(2, sheet.max_column + 1):
        if sheet.cell(row=4, column=c).value == 2:
            col_2y = c
            break

    if col_2y is None:
        raise RuntimeError("Could not find 2-year maturity column")

    # Collect all valid rows (most recent first)
    rows = []
    for r in range(sheet.max_row, 1, -1):
        d = sheet.cell(row=r, column=1).value
        v = sheet.cell(row=r, column=col_2y).value

        if d is None or v is None:
            continue
        if isinstance(d, str) and "maturity" in d.lower():
            continue

        if hasattr(d, "strftime"):
            date_str = d.strftime("%Y-%m-%d")
        else:
            date_str = str(d)[:10]

        rows.append({"date": date_str, "yield_2y": float(v)})

    return rows


def main():
    # 1. Get all available days from the BoE file
    boe_rows = fetch_latest_from_boe()
    if not boe_rows:
        raise RuntimeError("Could not find any valid data rows")

    # 2. Load existing history
    history = load_history()

    # 3. Merge (Boe data wins for the same date)
    by_date = {item["date"]: item for item in history}
    for row in boe_rows:
        by_date[row["date"]] = row

    # 4. Sort by date and keep only the last 16 days
    merged = sorted(by_date.values(), key=lambda x: x["date"])
    history = merged[-MAX_DAYS:]

    # 5. Save
    save_history(history)
    print(f"Saved {len(history)} days of data (latest: {history[-1]})")


if __name__ == "__main__":
    main()
